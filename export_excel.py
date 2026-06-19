from io import BytesIO
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import database as db


SHEET_ORDER = [
    "Project_Budget_Master",
    "Daily_Worklog",
    "Weekly_Summary_Input",
    "Weekly_Report_Log",
    "Project_Hour_Summary",
    "Staff_Weekly_Summary",
    "Missing_Weekly_Summary",
    "Dashboard_Source",
    "Master_Lists",
    "User_Master",
    "Role_Permission",
    "Audit_Log",
]


def _scope_raw_table(table_name, user):
    raw = db.get_table(table_name)
    if user is None or user["Role"] == "Admin":
        return raw

    role = user["Role"]

    if table_name == "daily_worklog":
        if role == "Manager":
            return raw[raw["Department"] == user["Department"]]
        if role == "Staff":
            return raw[raw["Staff_Name"] == user["Staff_Name"]]
        if role == "PD/PM":
            return raw[raw["Cost_Tracking_ID"].isin(db.get_owned_cost_ids(user["Staff_Name"]))]
        return raw.iloc[0:0]

    if table_name == "weekly_summary_input":
        if role == "Staff":
            return raw[raw["Staff_Name"] == user["Staff_Name"]]
        if role == "PD/PM":
            return raw[raw["Cost_Tracking_ID"].isin(db.get_owned_cost_ids(user["Staff_Name"]))]
        if role == "Manager":
            staff_df = db.get_staff()
            names = staff_df[staff_df["Department"] == user["Department"]]["Staff_Name"].tolist()
            return raw[raw["Staff_Name"].isin(names)]
        return raw.iloc[0:0]

    if table_name == "weekly_report_log":
        return db.scope_weekly_df(raw, user)

    if table_name == "project_budget_master":
        return db.scope_project_df(raw, user)

    if table_name in ["user_master", "role_permission", "audit_log"]:
        return raw.iloc[0:0]

    return raw


def get_export_frames(user=None):
    db.rebuild_weekly_report_log()
    admin = user is not None and user["Role"] == "Admin"
    return {
        "Project_Budget_Master": _scope_raw_table("project_budget_master", user),
        "Daily_Worklog": _scope_raw_table("daily_worklog", user),
        "Weekly_Summary_Input": _scope_raw_table("weekly_summary_input", user),
        "Weekly_Report_Log": _scope_raw_table("weekly_report_log", user),
        "Project_Hour_Summary": db.get_project_hour_summary(user=user),
        "Staff_Weekly_Summary": db.get_staff_weekly_summary(user=user),
        "Missing_Weekly_Summary": db.get_missing_weekly_summary(user=user),
        "Dashboard_Source": db.get_project_hour_summary(user=user),
        "Master_Lists": db.get_table("master_lists") if admin else pd.DataFrame(),
        "User_Master": db.get_table("user_master") if admin else pd.DataFrame(),
        "Role_Permission": db.get_table("role_permission") if admin else pd.DataFrame(),
        "Audit_Log": db.get_table("audit_log") if admin else pd.DataFrame(),
    }


def _write_dataframe(ws, df):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    if df is None or df.empty:
        ws.append(["No Data"])
        ws["A1"].font = header_font
        return

    ws.append(list(df.columns))
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for _, row in df.iterrows():
        ws.append([row[col] for col in df.columns])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    percent_cols = {"Budget_Burn_Rate", "Sales_Estimate_Burn_Rate"}
    numeric_keywords = ["Hours", "Rate", "Count", "Budget", "Estimated", "Remaining"]

    for col_idx, column_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        if column_name in percent_cols:
            for row_idx in range(2, ws.max_row + 1):
                ws[f"{col_letter}{row_idx}"].number_format = "0.0%"
        elif any(k in str(column_name) for k in numeric_keywords):
            for row_idx in range(2, ws.max_row + 1):
                ws[f"{col_letter}{row_idx}"].number_format = "0.0"

        max_len = len(str(column_name))
        for row_idx in range(2, min(ws.max_row, 200) + 1):
            value = ws[f"{col_letter}{row_idx}"].value
            if value is not None:
                max_len = max(max_len, min(len(str(value)), 60))
        ws.column_dimensions[col_letter].width = max(12, min(max_len + 2, 65))

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def export_full_workbook(user=None):
    frames = get_export_frames(user=user)
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name in SHEET_ORDER:
        ws = wb.create_sheet(sheet_name)
        _write_dataframe(ws, frames.get(sheet_name, pd.DataFrame()))

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    role_suffix = user["Role"].replace("/", "_") if user else "ALL"
    filename = f"GLOBAL_Weekly_Report_Demo_v3_{role_suffix}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return filename, output.getvalue()


def export_query_result(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Query_Result"
    _write_dataframe(ws, df)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"GLOBAL_Weekly_Report_Query_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return filename, output.getvalue()

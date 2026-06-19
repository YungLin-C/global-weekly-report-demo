import sqlite3
from pathlib import Path
from datetime import datetime, date
import random
import json
from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


DB_NAME = "global_weekly_report_demo.db"
DB_PATH = Path(__file__).parent / DB_NAME

ROLES = ["Admin", "Manager", "Staff", "PD/PM", "Viewer"]

PAGE_ORDER = [
    "預算設定",
    "日報輸入",
    "週報輸入",
    "自由彙整",
    "自動彙整 Dashboard",
    "缺漏週報",
    "匯出 Excel",
    "Master Data",
    "User Management",
    "Admin Data Maintenance",
    "Admin Config Import / Export",
]

STAFF_SEED = [
    ("小李", "機械設計", "Engineer", 1),
    ("張三", "電氣", "Engineer", 1),
    ("王四", "組立", "Engineer", 1),
    ("JOHN", "CS", "Engineer", 1),
    ("鈴木", "營業", "PD/PM", 1),
    ("大島", "機械設計", "Engineer", 1),
    ("MIKE", "電氣", "Engineer", 1),
    ("MERRY", "組立", "Engineer", 1),
    ("MAX", "CS", "Engineer", 1),
    ("JAME", "營業", "PD/PM", 1),
    ("大河", "機械設計", "Engineer", 1),
    ("吾郎", "電氣", "Engineer", 1),
    ("人傑", "組立", "Engineer", 1),
    ("大勇", "CS", "Engineer", 1),
]

MASTER_LIST_SEED = {
    "Department": ["機械設計", "電氣", "組立", "CS", "營業"],
    "Product_Line": ["OWLS-1800", "OWLS-2200", "OCSS-600", "OTFC-1800", "T"],
    "Platform_Line": ["機械設計", "電氣", "組立", "CS", "營業"],
    "Work_Category": ["定例會議", "資料彙整", "設計", "組立", "調適", "現場異常", "異常分析"],
    "Health": ["Green", "Yellow", "Red"],
    "Status": ["Active", "Hold", "Closed"],
    "Role": ROLES,
}

DEFAULT_USERS = [
    ("admin@demo.com", "小李", "機械設計", "Admin", 1),
    ("manager@demo.com", "張三", "電氣", "Manager", 1),
    ("staff@demo.com", "王四", "組立", "Staff", 1),
    ("pdpm@demo.com", "鈴木", "營業", "PD/PM", 1),
    ("viewer@demo.com", "JOHN", "CS", "Viewer", 1),
]

DEFAULT_COST_IDS = ["W0001", "W0002", "W0003", "W0004", "A0011", "A0012"]


def get_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def execute(sql, params=None):
    with get_connection() as conn:
        conn.execute(sql, params or [])
        conn.commit()


def query_df(sql, params=None):
    with get_connection() as conn:
        return pd.read_sql_query(sql, conn, params=params or [])


def now():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def get_current_week(input_date=None):
    if input_date is None:
        input_date = date.today()
    if isinstance(input_date, str):
        input_date = datetime.strptime(input_date[:10], "%Y-%m-%d").date()
    iso = input_date.isocalendar()
    return f"{iso.year}-W{iso.week:02d}"


def _column_exists(conn, table_name, column_name):
    rows = conn.execute(f"PRAGMA table_info({table_name})").fetchall()
    return any(row[1] == column_name for row in rows)


def _safe_add_column(conn, table_name, column_definition):
    column_name = column_definition.split()[0]
    if not _column_exists(conn, table_name, column_name):
        conn.execute(f"ALTER TABLE {table_name} ADD COLUMN {column_definition}")


def init_db():
    with get_connection() as conn:
        c = conn.cursor()

        c.execute("""
        CREATE TABLE IF NOT EXISTS staff_master (
            Staff_Name TEXT PRIMARY KEY,
            Department TEXT NOT NULL,
            Role TEXT,
            Active_Flag INTEGER DEFAULT 1
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS master_lists (
            List_Type TEXT NOT NULL,
            List_Value TEXT NOT NULL,
            Active_Flag INTEGER DEFAULT 1,
            PRIMARY KEY (List_Type, List_Value)
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS project_budget_master (
            Opportunity_ID TEXT,
            Cost_Tracking_ID TEXT PRIMARY KEY,
            Customer TEXT NOT NULL,
            Project_Name TEXT NOT NULL,
            Product_Line TEXT NOT NULL,
            Platform_Line TEXT NOT NULL,
            Budget_Hours REAL NOT NULL,
            Sales_Estimated_Hours REAL NOT NULL,
            Owner TEXT,
            Status TEXT NOT NULL DEFAULT 'Active',
            Created_At TEXT,
            Updated_At TEXT,
            Created_By TEXT,
            Updated_By TEXT
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS daily_worklog (
            Worklog_ID TEXT PRIMARY KEY,
            Work_Date TEXT NOT NULL,
            Report_Week TEXT NOT NULL,
            Staff_Name TEXT NOT NULL,
            Department TEXT NOT NULL,
            Cost_Tracking_ID TEXT NOT NULL,
            Work_Category TEXT NOT NULL,
            Hours REAL NOT NULL,
            Work_Content TEXT NOT NULL,
            Product_Line TEXT,
            Platform_Line TEXT,
            Created_At TEXT,
            Created_By TEXT
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS weekly_summary_input (
            Weekly_Input_ID TEXT PRIMARY KEY,
            Report_Week TEXT NOT NULL,
            Staff_Name TEXT NOT NULL,
            Cost_Tracking_ID TEXT NOT NULL,
            Weekly_Summary TEXT NOT NULL,
            Next_Week_Target TEXT NOT NULL,
            Health TEXT NOT NULL,
            Created_At TEXT,
            Updated_At TEXT,
            Created_By TEXT,
            Updated_By TEXT,
            UNIQUE (Report_Week, Staff_Name, Cost_Tracking_ID)
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS weekly_report_log (
            Report_Week TEXT NOT NULL,
            Staff_Name TEXT NOT NULL,
            Department TEXT NOT NULL,
            Cost_Tracking_ID TEXT NOT NULL,
            Customer TEXT,
            Project_Name TEXT,
            Product_Line TEXT,
            Platform_Line TEXT,
            Weekly_Total_Hours REAL,
            Daily_Work_Detail TEXT,
            Work_Category_Summary TEXT,
            Weekly_Summary TEXT,
            Next_Week_Target TEXT,
            Health TEXT,
            Submit_Status TEXT,
            Updated_At TEXT,
            PRIMARY KEY (Report_Week, Staff_Name, Cost_Tracking_ID)
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS user_master (
            User_Email TEXT PRIMARY KEY,
            Staff_Name TEXT,
            Department TEXT,
            Role TEXT NOT NULL,
            Active_Flag INTEGER DEFAULT 1,
            Created_At TEXT,
            Updated_At TEXT
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS role_permission (
            Role TEXT NOT NULL,
            Page_Name TEXT NOT NULL,
            Can_View INTEGER DEFAULT 0,
            Can_Edit INTEGER DEFAULT 0,
            Can_Export INTEGER DEFAULT 0,
            PRIMARY KEY (Role, Page_Name)
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS audit_log (
            Audit_ID TEXT PRIMARY KEY,
            User_Email TEXT,
            Action_Type TEXT,
            Table_Name TEXT,
            Record_Key TEXT,
            Old_Value TEXT,
            New_Value TEXT,
            Created_At TEXT
        )
        """)

        for table, columns in {
            "project_budget_master": ["Created_By TEXT", "Updated_By TEXT"],
            "daily_worklog": ["Created_By TEXT"],
            "weekly_summary_input": ["Created_By TEXT", "Updated_By TEXT"],
        }.items():
            for col in columns:
                _safe_add_column(conn, table, col)

        conn.commit()

    seed_master_data()
    seed_users_and_permissions()
    seed_project_budget_master()


def seed_master_data():
    with get_connection() as conn:
        c = conn.cursor()

        if c.execute("SELECT COUNT(*) FROM staff_master").fetchone()[0] == 0:
            c.executemany("""
                INSERT INTO staff_master (Staff_Name, Department, Role, Active_Flag)
                VALUES (?, ?, ?, ?)
            """, STAFF_SEED)

        c.execute("UPDATE staff_master SET Role='PD/PM' WHERE Role='Sales'")

        for list_type, values in MASTER_LIST_SEED.items():
            for value in values:
                c.execute("""
                    INSERT OR IGNORE INTO master_lists (List_Type, List_Value, Active_Flag)
                    VALUES (?, ?, 1)
                """, [list_type, value])

        c.execute("UPDATE master_lists SET List_Value='PD/PM' WHERE List_Type='Role' AND List_Value='Sales'")
        conn.commit()


def seed_users_and_permissions():
    n = now()
    with get_connection() as conn:
        c = conn.cursor()

        for email, staff, dept, role, active in DEFAULT_USERS:
            c.execute("""
                INSERT OR IGNORE INTO user_master
                (User_Email, Staff_Name, Department, Role, Active_Flag, Created_At, Updated_At)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, [email, staff, dept, role, active, n, n])

        if c.execute("SELECT COUNT(*) FROM role_permission").fetchone()[0] == 0:
            reset_default_permissions(conn)

        conn.commit()


def reset_default_permissions(conn=None):
    own_conn = conn is None
    if own_conn:
        conn = get_connection()
    c = conn.cursor()
    c.execute("DELETE FROM role_permission")

    permissions = {
        "Admin": {
            "預算設定": (1, 1, 1),
            "日報輸入": (1, 1, 0),
            "週報輸入": (1, 1, 0),
            "自由彙整": (1, 0, 1),
            "自動彙整 Dashboard": (1, 0, 1),
            "缺漏週報": (1, 0, 1),
            "匯出 Excel": (1, 0, 1),
            "Master Data": (1, 1, 1),
            "User Management": (1, 1, 1),
            "Admin Data Maintenance": (1, 1, 1),
            "Admin Config Import / Export": (1, 1, 1),
        },
        "Manager": {
            "日報輸入": (1, 1, 0),
            "週報輸入": (1, 1, 0),
            "自由彙整": (1, 0, 1),
            "自動彙整 Dashboard": (1, 0, 1),
            "缺漏週報": (1, 0, 1),
            "匯出 Excel": (1, 0, 1),
        },
        "Staff": {
            "日報輸入": (1, 1, 0),
            "週報輸入": (1, 1, 0),
            "自由彙整": (1, 0, 1),
            "自動彙整 Dashboard": (1, 0, 0),
            "缺漏週報": (1, 0, 0),
            "匯出 Excel": (1, 0, 1),
        },
        "PD/PM": {
            "預算設定": (1, 1, 1),
            "日報輸入": (1, 1, 0),
            "週報輸入": (1, 1, 0),
            "自由彙整": (1, 0, 1),
            "自動彙整 Dashboard": (1, 0, 1),
            "缺漏週報": (1, 0, 1),
            "匯出 Excel": (1, 0, 1),
        },
        "Viewer": {
            "自動彙整 Dashboard": (1, 0, 0),
        },
    }

    for role, pages in permissions.items():
        for page, flags in pages.items():
            c.execute("""
                INSERT INTO role_permission (Role, Page_Name, Can_View, Can_Edit, Can_Export)
                VALUES (?, ?, ?, ?, ?)
            """, [role, page, flags[0], flags[1], flags[2]])
    if own_conn:
        conn.commit()
        conn.close()


def seed_project_budget_master():
    with get_connection() as conn:
        c = conn.cursor()
        if c.execute("SELECT COUNT(*) FROM project_budget_master").fetchone()[0] > 0:
            return

        customers = [f"DEMO Customer {x}" for x in ["A", "B", "C", "D", "E", "F"]]
        product_lines = get_master_values("Product_Line")
        platform_lines = get_master_values("Platform_Line")
        pdpm_names = query_df("SELECT Staff_Name FROM staff_master WHERE Role='PD/PM' AND Active_Flag=1")["Staff_Name"].tolist()
        staff_names = query_df("SELECT Staff_Name FROM staff_master WHERE Active_Flag=1")["Staff_Name"].tolist()
        owners = pdpm_names or staff_names
        budgets = [50, 100, 150, 200, 300]
        n = now()

        rows = []
        for idx, cost_id in enumerate(DEFAULT_COST_IDS, 1):
            budget = random.choice(budgets)
            rows.append([
                f"OPP-2026-{idx:03d}",
                cost_id,
                random.choice(customers),
                f"DEMO Project {cost_id}",
                random.choice(product_lines),
                random.choice(platform_lines),
                budget,
                round(budget * random.uniform(0.8, 1.2), 1),
                random.choice(owners),
                "Active",
                n,
                n,
                "system",
                "system",
            ])

        c.executemany("""
            INSERT INTO project_budget_master (
                Opportunity_ID, Cost_Tracking_ID, Customer, Project_Name, Product_Line,
                Platform_Line, Budget_Hours, Sales_Estimated_Hours, Owner, Status,
                Created_At, Updated_At, Created_By, Updated_By
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, rows)
        conn.commit()


def generate_id(prefix, table_name, id_col):
    year = datetime.now().year
    pattern = f"{prefix}-{year}-%"
    with get_connection() as conn:
        row = conn.execute(
            f"SELECT {id_col} FROM {table_name} WHERE {id_col} LIKE ? ORDER BY {id_col} DESC LIMIT 1",
            [pattern],
        ).fetchone()
    if row is None:
        next_num = 1
    else:
        try:
            next_num = int(row[0].split("-")[-1]) + 1
        except Exception:
            next_num = 1
    return f"{prefix}-{year}-{next_num:06d}"


def audit(action_type, table_name, record_key, old_value="", new_value="", user_email="system"):
    audit_id = generate_id("AUD", "audit_log", "Audit_ID")
    execute("""
        INSERT INTO audit_log (Audit_ID, User_Email, Action_Type, Table_Name, Record_Key, Old_Value, New_Value, Created_At)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, [
        audit_id,
        user_email,
        action_type,
        table_name,
        str(record_key),
        json.dumps(old_value, ensure_ascii=False, default=str) if not isinstance(old_value, str) else old_value,
        json.dumps(new_value, ensure_ascii=False, default=str) if not isinstance(new_value, str) else new_value,
        now(),
    ])


def get_master_values(list_type):
    df = query_df("""
        SELECT List_Value FROM master_lists
        WHERE List_Type=? AND Active_Flag=1
        ORDER BY List_Value
    """, [list_type])
    return df["List_Value"].tolist() if not df.empty else []


def get_table(table_name):
    allowed = {
        "project_budget_master", "staff_master", "daily_worklog", "weekly_summary_input",
        "weekly_report_log", "master_lists", "user_master", "role_permission", "audit_log"
    }
    if table_name not in allowed:
        raise ValueError("Unsupported table")
    return query_df(f"SELECT * FROM {table_name}")


def get_staff(active_only=True):
    sql = "SELECT * FROM staff_master"
    if active_only:
        sql += " WHERE Active_Flag=1"
    sql += " ORDER BY Staff_Name"
    return query_df(sql)


def get_users(active_only=True):
    sql = "SELECT * FROM user_master"
    if active_only:
        sql += " WHERE Active_Flag=1"
    sql += " ORDER BY Role, User_Email"
    return query_df(sql)


def get_user_by_email(email):
    df = query_df("SELECT * FROM user_master WHERE User_Email=? AND Active_Flag=1", [email])
    return None if df.empty else df.iloc[0].to_dict()


def get_allowed_pages(role):
    df = query_df("""
        SELECT Page_Name FROM role_permission
        WHERE Role=? AND Can_View=1
    """, [role])
    pages = df["Page_Name"].tolist() if not df.empty else []
    return [p for p in PAGE_ORDER if p in pages]


def get_permission(role, page_name):
    df = query_df("""
        SELECT Can_View, Can_Edit, Can_Export
        FROM role_permission
        WHERE Role=? AND Page_Name=?
    """, [role, page_name])
    if df.empty:
        return {"Can_View": 0, "Can_Edit": 0, "Can_Export": 0}
    row = df.iloc[0]
    return {
        "Can_View": int(row["Can_View"]),
        "Can_Edit": int(row["Can_Edit"]),
        "Can_Export": int(row["Can_Export"]),
    }


def get_owned_cost_ids(staff_name):
    df = query_df("SELECT Cost_Tracking_ID FROM project_budget_master WHERE Owner=? ORDER BY Cost_Tracking_ID", [staff_name])
    return df["Cost_Tracking_ID"].tolist() if not df.empty else []


def scope_project_df(df, user):
    if df.empty or not user:
        return df
    role = user["Role"]
    if role in ["Admin", "Viewer"]:
        return df
    if role == "PD/PM":
        return df[df["Owner"] == user["Staff_Name"]] if "Owner" in df.columns else df.iloc[0:0]
    return df


def scope_weekly_df(df, user):
    if df.empty or not user:
        return df
    role = user["Role"]
    if role in ["Admin", "Viewer"]:
        return df
    if role == "Manager" and "Department" in df.columns:
        return df[df["Department"] == user["Department"]]
    if role == "Staff" and "Staff_Name" in df.columns:
        return df[df["Staff_Name"] == user["Staff_Name"]]
    if role == "PD/PM" and "Cost_Tracking_ID" in df.columns:
        return df[df["Cost_Tracking_ID"].isin(get_owned_cost_ids(user["Staff_Name"]))]
    return df.iloc[0:0]


def get_project_budget_for_display(user=None):
    return scope_project_df(get_table("project_budget_master"), user)


def get_active_cost_ids(user=None):
    df = query_df("SELECT Cost_Tracking_ID, Owner FROM project_budget_master WHERE Status='Active' ORDER BY Cost_Tracking_ID")
    if user is not None and user["Role"] == "PD/PM":
        df = df[df["Owner"] == user["Staff_Name"]]
    return df["Cost_Tracking_ID"].tolist() if not df.empty else []


def upsert_project_budget(record, user=None):
    required = ["Cost_Tracking_ID", "Customer", "Project_Name", "Product_Line", "Platform_Line", "Status"]
    for col in required:
        if not record.get(col):
            raise ValueError(f"{col} 不可為空。")
    if float(record.get("Budget_Hours") or 0) <= 0:
        raise ValueError("Budget_Hours 必須大於 0。")

    sales_hours = record.get("Sales_Estimated_Hours") or record["Budget_Hours"]
    user_email = user["User_Email"] if user else "system"
    n = now()

    with get_connection() as conn:
        old = conn.execute("SELECT * FROM project_budget_master WHERE Cost_Tracking_ID=?", [record["Cost_Tracking_ID"]]).fetchone()

        if user and user["Role"] == "PD/PM":
            if old and old["Owner"] != user["Staff_Name"]:
                raise PermissionError("PD/PM 只能修改自己 Owner 的 Cost_Tracking_ID。")
            record["Owner"] = user["Staff_Name"]

        if old:
            conn.execute("""
                UPDATE project_budget_master
                SET Opportunity_ID=?, Customer=?, Project_Name=?, Product_Line=?, Platform_Line=?,
                    Budget_Hours=?, Sales_Estimated_Hours=?, Owner=?, Status=?, Updated_At=?, Updated_By=?
                WHERE Cost_Tracking_ID=?
            """, [
                record.get("Opportunity_ID"), record["Customer"], record["Project_Name"],
                record["Product_Line"], record["Platform_Line"], float(record["Budget_Hours"]),
                float(sales_hours), record.get("Owner"), record["Status"], n, user_email,
                record["Cost_Tracking_ID"],
            ])
            action = "UPDATE"
        else:
            conn.execute("""
                INSERT INTO project_budget_master (
                    Opportunity_ID, Cost_Tracking_ID, Customer, Project_Name, Product_Line,
                    Platform_Line, Budget_Hours, Sales_Estimated_Hours, Owner, Status,
                    Created_At, Updated_At, Created_By, Updated_By
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, [
                record.get("Opportunity_ID"), record["Cost_Tracking_ID"], record["Customer"],
                record["Project_Name"], record["Product_Line"], record["Platform_Line"],
                float(record["Budget_Hours"]), float(sales_hours), record.get("Owner"),
                record["Status"], n, n, user_email, user_email,
            ])
            action = "INSERT"
        conn.commit()

    audit(action, "project_budget_master", record["Cost_Tracking_ID"], dict(old) if old else "", record, user_email)


def add_daily_worklog(work_date, staff_name, cost_tracking_id, work_category, hours, content, user=None):
    if user:
        if user["Role"] == "Staff" and staff_name != user["Staff_Name"]:
            raise PermissionError("Staff 只能替自己輸入日報。")
        if user["Role"] == "PD/PM" and cost_tracking_id not in get_owned_cost_ids(user["Staff_Name"]):
            raise PermissionError("PD/PM 只能輸入自己 Owner 的 Cost_Tracking_ID。")
        if user["Role"] == "Manager":
            sdf = query_df("SELECT Department FROM staff_master WHERE Staff_Name=?", [staff_name])
            if not sdf.empty and sdf.iloc[0]["Department"] != user["Department"]:
                raise PermissionError("Manager 只能替本部門人員輸入。")

    if float(hours) < 0.5 or float(hours) > 24:
        raise ValueError("工時必須介於 0.5 到 24。")
    if not content or not content.strip():
        raise ValueError("Work_Content 不可為空。")
    if len(content) > 100:
        raise ValueError("Work_Content 限制 100 字以內。")

    sdf = query_df("SELECT * FROM staff_master WHERE Staff_Name=?", [staff_name])
    pdf = query_df("SELECT * FROM project_budget_master WHERE Cost_Tracking_ID=?", [cost_tracking_id])
    if sdf.empty:
        raise ValueError("人員不存在。")
    if pdf.empty:
        raise ValueError("工番號不存在。")

    work_date_str = work_date.strftime("%Y-%m-%d") if hasattr(work_date, "strftime") else str(work_date)
    report_week = get_current_week(work_date_str)
    worklog_id = generate_id("WL", "daily_worklog", "Worklog_ID")
    user_email = user["User_Email"] if user else "system"

    execute("""
        INSERT INTO daily_worklog (
            Worklog_ID, Work_Date, Report_Week, Staff_Name, Department, Cost_Tracking_ID,
            Work_Category, Hours, Work_Content, Product_Line, Platform_Line, Created_At, Created_By
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, [
        worklog_id, work_date_str, report_week, staff_name, sdf.iloc[0]["Department"],
        cost_tracking_id, work_category, float(hours), content.strip(),
        pdf.iloc[0]["Product_Line"], pdf.iloc[0]["Platform_Line"], now(), user_email,
    ])

    rebuild_weekly_report_log()
    audit("INSERT", "daily_worklog", worklog_id, "", {"Hours": hours, "Cost_Tracking_ID": cost_tracking_id}, user_email)
    return worklog_id


def upsert_weekly_summary(report_week, staff_name, cost_tracking_id, summary, target, health, user=None):
    if user:
        if user["Role"] == "Staff" and staff_name != user["Staff_Name"]:
            raise PermissionError("Staff 只能替自己輸入週報。")
        if user["Role"] == "PD/PM" and cost_tracking_id not in get_owned_cost_ids(user["Staff_Name"]):
            raise PermissionError("PD/PM 只能管理自己 Owner 的 Cost_Tracking_ID。")
        if user["Role"] == "Manager":
            sdf = query_df("SELECT Department FROM staff_master WHERE Staff_Name=?", [staff_name])
            if not sdf.empty and sdf.iloc[0]["Department"] != user["Department"]:
                raise PermissionError("Manager 只能替本部門人員輸入週報。")

    if not summary or not summary.strip():
        raise ValueError("Weekly_Summary 不可為空。")
    if not target or not target.strip():
        raise ValueError("Next_Week_Target 不可為空。")
    if len(summary) > 500 or len(target) > 500:
        raise ValueError("週報欄位限制 500 字以內。")

    user_email = user["User_Email"] if user else "system"
    n = now()

    with get_connection() as conn:
        old = conn.execute("""
            SELECT * FROM weekly_summary_input
            WHERE Report_Week=? AND Staff_Name=? AND Cost_Tracking_ID=?
        """, [report_week, staff_name, cost_tracking_id]).fetchone()

        if old:
            wid = old["Weekly_Input_ID"]
            conn.execute("""
                UPDATE weekly_summary_input
                SET Weekly_Summary=?, Next_Week_Target=?, Health=?, Updated_At=?, Updated_By=?
                WHERE Weekly_Input_ID=?
            """, [summary.strip(), target.strip(), health, n, user_email, wid])
            action = "UPDATE"
        else:
            wid = generate_id("WSI", "weekly_summary_input", "Weekly_Input_ID")
            conn.execute("""
                INSERT INTO weekly_summary_input (
                    Weekly_Input_ID, Report_Week, Staff_Name, Cost_Tracking_ID,
                    Weekly_Summary, Next_Week_Target, Health, Created_At, Updated_At, Created_By, Updated_By
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, [
                wid, report_week, staff_name, cost_tracking_id, summary.strip(), target.strip(),
                health, n, n, user_email, user_email,
            ])
            action = "INSERT"
        conn.commit()

    rebuild_weekly_report_log()
    audit(action, "weekly_summary_input", wid, dict(old) if old else "", {
        "Report_Week": report_week,
        "Staff_Name": staff_name,
        "Cost_Tracking_ID": cost_tracking_id,
    }, user_email)
    return wid


def rebuild_weekly_report_log():
    execute("DELETE FROM weekly_report_log")

    base = query_df("""
        SELECT
            dw.Report_Week, dw.Staff_Name, dw.Department, dw.Cost_Tracking_ID,
            pb.Customer, pb.Project_Name, pb.Product_Line, pb.Platform_Line,
            SUM(dw.Hours) AS Weekly_Total_Hours
        FROM daily_worklog dw
        LEFT JOIN project_budget_master pb ON dw.Cost_Tracking_ID = pb.Cost_Tracking_ID
        GROUP BY dw.Report_Week, dw.Staff_Name, dw.Department, dw.Cost_Tracking_ID,
                 pb.Customer, pb.Project_Name, pb.Product_Line, pb.Platform_Line
    """)

    with get_connection() as conn:
        for _, row in base.iterrows():
            details = query_df("""
                SELECT Work_Date, Work_Category, Hours, Work_Content
                FROM daily_worklog
                WHERE Report_Week=? AND Staff_Name=? AND Cost_Tracking_ID=?
                ORDER BY Work_Date, Created_At
            """, [row["Report_Week"], row["Staff_Name"], row["Cost_Tracking_ID"]])

            detail_text = "\n".join([
                f'{d["Work_Date"]} | {d["Work_Category"]} | {d["Hours"]}h | {d["Work_Content"]}'
                for _, d in details.iterrows()
            ])

            cats = query_df("""
                SELECT Work_Category, SUM(Hours) AS Hours
                FROM daily_worklog
                WHERE Report_Week=? AND Staff_Name=? AND Cost_Tracking_ID=?
                GROUP BY Work_Category
                ORDER BY Work_Category
            """, [row["Report_Week"], row["Staff_Name"], row["Cost_Tracking_ID"]])
            cat_text = " / ".join([f'{c["Work_Category"]}: {c["Hours"]}h' for _, c in cats.iterrows()])

            wsi = query_df("""
                SELECT Weekly_Summary, Next_Week_Target, Health
                FROM weekly_summary_input
                WHERE Report_Week=? AND Staff_Name=? AND Cost_Tracking_ID=?
            """, [row["Report_Week"], row["Staff_Name"], row["Cost_Tracking_ID"]])

            if wsi.empty:
                weekly_summary = ""
                next_target = ""
                health = ""
                status = "Missing Summary"
            else:
                weekly_summary = wsi.iloc[0]["Weekly_Summary"]
                next_target = wsi.iloc[0]["Next_Week_Target"]
                health = wsi.iloc[0]["Health"]
                status = "Submitted"

            conn.execute("""
                INSERT INTO weekly_report_log (
                    Report_Week, Staff_Name, Department, Cost_Tracking_ID,
                    Customer, Project_Name, Product_Line, Platform_Line, Weekly_Total_Hours,
                    Daily_Work_Detail, Work_Category_Summary, Weekly_Summary, Next_Week_Target,
                    Health, Submit_Status, Updated_At
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, [
                row["Report_Week"], row["Staff_Name"], row["Department"], row["Cost_Tracking_ID"],
                row["Customer"], row["Project_Name"], row["Product_Line"], row["Platform_Line"],
                float(row["Weekly_Total_Hours"] or 0), detail_text, cat_text,
                weekly_summary, next_target, health, status, now(),
            ])

        conn.commit()


def get_week_options():
    df = query_df("""
        SELECT Report_Week FROM daily_worklog
        UNION
        SELECT Report_Week FROM weekly_summary_input
        ORDER BY Report_Week DESC
    """)
    options = df["Report_Week"].tolist() if not df.empty else []
    current = get_current_week()
    if current not in options:
        options.insert(0, current)
    return options


def get_cost_ids_for_staff_week(staff_name, report_week, user=None):
    df = query_df("""
        SELECT DISTINCT dw.Cost_Tracking_ID, pb.Owner
        FROM daily_worklog dw
        LEFT JOIN project_budget_master pb ON dw.Cost_Tracking_ID = pb.Cost_Tracking_ID
        WHERE dw.Staff_Name=? AND dw.Report_Week=?
        ORDER BY dw.Cost_Tracking_ID
    """, [staff_name, report_week])
    if user is not None and user["Role"] == "PD/PM":
        df = df[df["Owner"] == user["Staff_Name"]]
    ids = df["Cost_Tracking_ID"].tolist() if not df.empty else []
    return ids or get_active_cost_ids(user)


def get_weekly_report_filtered(start_week=None, end_week=None, department=None, staff_name=None, cost_tracking_id=None, user=None):
    where = []
    params = []
    if start_week:
        where.append("Report_Week >= ?")
        params.append(start_week)
    if end_week:
        where.append("Report_Week <= ?")
        params.append(end_week)
    if department and department != "ALL":
        where.append("Department = ?")
        params.append(department)
    if staff_name and staff_name != "ALL":
        where.append("Staff_Name = ?")
        params.append(staff_name)
    if cost_tracking_id and cost_tracking_id != "ALL":
        where.append("Cost_Tracking_ID = ?")
        params.append(cost_tracking_id)

    sql_where = "WHERE " + " AND ".join(where) if where else ""
    df = query_df(f"""
        SELECT Report_Week, Staff_Name, Department, Cost_Tracking_ID, Customer, Project_Name,
               Product_Line, Platform_Line, Weekly_Total_Hours, Weekly_Summary,
               Next_Week_Target, Health, Submit_Status
        FROM weekly_report_log
        {sql_where}
        ORDER BY Report_Week DESC, Staff_Name, Cost_Tracking_ID
    """, params)
    return scope_weekly_df(df, user) if user else df


def get_project_hour_summary(report_week=None, user=None):
    current_week = report_week or get_current_week()
    df = query_df("""
        SELECT
            pb.Cost_Tracking_ID, pb.Customer, pb.Project_Name, pb.Product_Line, pb.Platform_Line,
            pb.Budget_Hours, pb.Sales_Estimated_Hours,
            COALESCE(SUM(CASE WHEN dw.Report_Week=? THEN dw.Hours ELSE 0 END), 0) AS Weekly_Hours,
            COALESCE(SUM(dw.Hours), 0) AS Cumulative_Hours,
            pb.Owner, pb.Status
        FROM project_budget_master pb
        LEFT JOIN daily_worklog dw ON pb.Cost_Tracking_ID = dw.Cost_Tracking_ID
        GROUP BY pb.Cost_Tracking_ID, pb.Customer, pb.Project_Name, pb.Product_Line,
                 pb.Platform_Line, pb.Budget_Hours, pb.Sales_Estimated_Hours, pb.Owner, pb.Status
        ORDER BY pb.Cost_Tracking_ID
    """, [current_week])
    df = scope_project_df(df, user) if user else df
    if df.empty:
        return df
    df["Remaining_Hours"] = df["Budget_Hours"] - df["Cumulative_Hours"]
    df["Budget_Burn_Rate"] = df.apply(lambda r: r["Cumulative_Hours"] / r["Budget_Hours"] if r["Budget_Hours"] else 0, axis=1)
    df["Sales_Estimate_Burn_Rate"] = df.apply(lambda r: r["Cumulative_Hours"] / r["Sales_Estimated_Hours"] if r["Sales_Estimated_Hours"] else 0, axis=1)
    df["Budget_Health"] = df["Budget_Burn_Rate"].apply(get_budget_health)
    return df


def get_budget_health(rate):
    if rate < 0.5:
        return "Green"
    if rate <= 1.0:
        return "Yellow"
    if rate <= 1.5:
        return "Red"
    if rate <= 2.0:
        return "Critical"
    return "Overrun"


def get_staff_weekly_summary(user=None):
    df = query_df("""
        SELECT Report_Week, Staff_Name, Department,
               SUM(Weekly_Total_Hours) AS Total_Weekly_Hours,
               COUNT(DISTINCT Cost_Tracking_ID) AS Number_of_Cost_IDs,
               SUM(CASE WHEN Submit_Status='Missing Summary' THEN 1 ELSE 0 END) AS Missing_Summary_Count
        FROM weekly_report_log
        GROUP BY Report_Week, Staff_Name, Department
        ORDER BY Report_Week DESC, Staff_Name
    """)
    return scope_weekly_df(df, user) if user else df


def get_missing_weekly_summary(user=None):
    df = query_df("""
        SELECT Report_Week, Staff_Name, Department, Cost_Tracking_ID, Customer, Project_Name,
               Weekly_Total_Hours, 'Weekly Summary' AS Missing_Item, 'Not Sent' AS Reminder_Status
        FROM weekly_report_log
        WHERE Submit_Status='Missing Summary'
        ORDER BY Report_Week DESC, Staff_Name, Cost_Tracking_ID
    """)
    return scope_weekly_df(df, user) if user else df


def _admin_only(user):
    if not user or user.get("Role") != "Admin":
        raise PermissionError("只有 Admin 可以執行此操作。")


def upsert_staff(staff_name, department, role="Engineer", active_flag=1, user=None):
    _admin_only(user)
    if not staff_name:
        raise ValueError("Staff_Name 不可為空。")
    old = query_df("SELECT * FROM staff_master WHERE Staff_Name=?", [staff_name])
    execute("""
        INSERT INTO staff_master (Staff_Name, Department, Role, Active_Flag)
        VALUES (?, ?, ?, ?)
        ON CONFLICT(Staff_Name)
        DO UPDATE SET Department=excluded.Department, Role=excluded.Role, Active_Flag=excluded.Active_Flag
    """, [staff_name.strip(), department, role, int(active_flag)])
    audit("ADMIN_UPSERT", "staff_master", staff_name, old.iloc[0].to_dict() if not old.empty else "", {
        "Department": department, "Role": role, "Active_Flag": int(active_flag)
    }, user["User_Email"])


def update_staff_master(staff_name, department, role, active_flag, user=None):
    return upsert_staff(staff_name, department, role, active_flag, user)


def delete_staff_master(staff_name, user=None):
    _admin_only(user)
    old = query_df("SELECT * FROM staff_master WHERE Staff_Name=?", [staff_name])
    if old.empty:
        raise ValueError("找不到 Staff_Name。")
    refs = []
    for table, col in [
        ("daily_worklog", "Staff_Name"),
        ("weekly_summary_input", "Staff_Name"),
        ("project_budget_master", "Owner"),
        ("user_master", "Staff_Name"),
    ]:
        if not query_df(f"SELECT 1 FROM {table} WHERE {col}=? LIMIT 1", [staff_name]).empty:
            refs.append(f"{table}.{col}")
    if refs:
        raise ValueError("此人員已被資料引用，請改 Active_Flag=0。引用：" + ", ".join(refs))
    execute("DELETE FROM staff_master WHERE Staff_Name=?", [staff_name])
    audit("ADMIN_DELETE", "staff_master", staff_name, old.iloc[0].to_dict(), "", user["User_Email"])


def upsert_user(user_email, staff_name, department, role, active_flag=1, user=None):
    _admin_only(user)
    if not user_email:
        raise ValueError("User_Email 不可為空。")
    if role not in ROLES:
        raise ValueError("Role 必須是 Admin / Manager / Staff / PD/PM / Viewer。")
    old = query_df("SELECT * FROM user_master WHERE User_Email=?", [user_email])
    n = now()
    execute("""
        INSERT INTO user_master (User_Email, Staff_Name, Department, Role, Active_Flag, Created_At, Updated_At)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(User_Email)
        DO UPDATE SET Staff_Name=excluded.Staff_Name,
                      Department=excluded.Department,
                      Role=excluded.Role,
                      Active_Flag=excluded.Active_Flag,
                      Updated_At=excluded.Updated_At
    """, [user_email.strip(), staff_name, department, role, int(active_flag), n, n])
    audit("ADMIN_UPSERT", "user_master", user_email, old.iloc[0].to_dict() if not old.empty else "", {
        "Staff_Name": staff_name, "Department": department, "Role": role, "Active_Flag": int(active_flag)
    }, user["User_Email"])


def delete_user(user_email, user=None):
    _admin_only(user)
    if user_email == user.get("User_Email"):
        raise ValueError("不可刪除目前登入中的自己。")
    old = query_df("SELECT * FROM user_master WHERE User_Email=?", [user_email])
    if old.empty:
        raise ValueError("找不到 User。")
    execute("DELETE FROM user_master WHERE User_Email=?", [user_email])
    audit("ADMIN_DELETE", "user_master", user_email, old.iloc[0].to_dict(), "", user["User_Email"])


def add_master_list_value(list_type, list_value, user=None):
    _admin_only(user)
    if not list_type or not list_value:
        raise ValueError("List_Type / List_Value 不可為空。")
    execute("""
        INSERT OR IGNORE INTO master_lists (List_Type, List_Value, Active_Flag)
        VALUES (?, ?, 1)
    """, [list_type, list_value.strip()])
    audit("ADMIN_INSERT_OR_IGNORE", "master_lists", f"{list_type}:{list_value}", "", "", user["User_Email"])


def update_master_list_value(old_list_type, old_list_value, new_list_type, new_list_value, active_flag=1, user=None):
    _admin_only(user)
    old = query_df("SELECT * FROM master_lists WHERE List_Type=? AND List_Value=?", [old_list_type, old_list_value])
    if old.empty:
        raise ValueError("找不到 Master List。")
    dup = query_df("""
        SELECT 1 FROM master_lists
        WHERE List_Type=? AND List_Value=? AND NOT (List_Type=? AND List_Value=?)
    """, [new_list_type, new_list_value.strip(), old_list_type, old_list_value])
    if not dup.empty:
        raise ValueError("修改後的 List_Type + List_Value 已存在。")
    execute("""
        UPDATE master_lists
        SET List_Type=?, List_Value=?, Active_Flag=?
        WHERE List_Type=? AND List_Value=?
    """, [new_list_type, new_list_value.strip(), int(active_flag), old_list_type, old_list_value])
    audit("ADMIN_UPDATE", "master_lists", f"{old_list_type}:{old_list_value}", old.iloc[0].to_dict(), {
        "List_Type": new_list_type, "List_Value": new_list_value.strip(), "Active_Flag": int(active_flag)
    }, user["User_Email"])


def delete_master_list_value(list_type, list_value, user=None):
    _admin_only(user)
    old = query_df("SELECT * FROM master_lists WHERE List_Type=? AND List_Value=?", [list_type, list_value])
    if old.empty:
        raise ValueError("找不到 Master List。")

    refs = []
    checks = {
        "Department": [("staff_master", "Department"), ("daily_worklog", "Department"), ("user_master", "Department")],
        "Product_Line": [("project_budget_master", "Product_Line"), ("daily_worklog", "Product_Line")],
        "Platform_Line": [("project_budget_master", "Platform_Line"), ("daily_worklog", "Platform_Line")],
        "Work_Category": [("daily_worklog", "Work_Category")],
        "Health": [("weekly_summary_input", "Health")],
        "Status": [("project_budget_master", "Status")],
        "Role": [("user_master", "Role")],
    }
    for table, col in checks.get(list_type, []):
        if not query_df(f"SELECT 1 FROM {table} WHERE {col}=? LIMIT 1", [list_value]).empty:
            refs.append(f"{table}.{col}")
    if refs:
        raise ValueError("此 Master Value 已被使用，請改 Active_Flag=0。引用：" + ", ".join(refs))

    execute("DELETE FROM master_lists WHERE List_Type=? AND List_Value=?", [list_type, list_value])
    audit("ADMIN_DELETE", "master_lists", f"{list_type}:{list_value}", old.iloc[0].to_dict(), "", user["User_Email"])


def get_daily_worklog_by_id(worklog_id):
    df = query_df("SELECT * FROM daily_worklog WHERE Worklog_ID=?", [worklog_id])
    return None if df.empty else df.iloc[0].to_dict()


def update_daily_worklog(worklog_id, work_date, staff_name, cost_tracking_id, work_category, hours, content, user=None):
    _admin_only(user)
    old = get_daily_worklog_by_id(worklog_id)
    if not old:
        raise ValueError("找不到 Worklog_ID。")
    sdf = query_df("SELECT * FROM staff_master WHERE Staff_Name=?", [staff_name])
    pdf = query_df("SELECT * FROM project_budget_master WHERE Cost_Tracking_ID=?", [cost_tracking_id])
    if sdf.empty:
        raise ValueError("人員不存在。")
    if pdf.empty:
        raise ValueError("工番號不存在。")
    if float(hours) < 0.5 or float(hours) > 24:
        raise ValueError("工時必須介於 0.5 到 24。")
    if not content or not content.strip():
        raise ValueError("Work_Content 不可為空。")
    if len(content) > 100:
        raise ValueError("Work_Content 限制 100 字以內。")

    wd = work_date.strftime("%Y-%m-%d") if hasattr(work_date, "strftime") else str(work_date)
    rw = get_current_week(wd)
    execute("""
        UPDATE daily_worklog
        SET Work_Date=?, Report_Week=?, Staff_Name=?, Department=?, Cost_Tracking_ID=?,
            Work_Category=?, Hours=?, Work_Content=?, Product_Line=?, Platform_Line=?
        WHERE Worklog_ID=?
    """, [
        wd, rw, staff_name, sdf.iloc[0]["Department"], cost_tracking_id, work_category,
        float(hours), content.strip(), pdf.iloc[0]["Product_Line"], pdf.iloc[0]["Platform_Line"], worklog_id
    ])
    rebuild_weekly_report_log()
    audit("ADMIN_UPDATE", "daily_worklog", worklog_id, old, {
        "Work_Date": wd, "Report_Week": rw, "Staff_Name": staff_name,
        "Cost_Tracking_ID": cost_tracking_id, "Hours": hours
    }, user["User_Email"])


def delete_daily_worklog(worklog_id, user=None):
    _admin_only(user)
    old = get_daily_worklog_by_id(worklog_id)
    if not old:
        raise ValueError("找不到 Worklog_ID。")
    execute("DELETE FROM daily_worklog WHERE Worklog_ID=?", [worklog_id])
    rebuild_weekly_report_log()
    audit("ADMIN_DELETE", "daily_worklog", worklog_id, old, "", user["User_Email"])


def get_weekly_summary_by_id(weekly_input_id):
    df = query_df("SELECT * FROM weekly_summary_input WHERE Weekly_Input_ID=?", [weekly_input_id])
    return None if df.empty else df.iloc[0].to_dict()


def update_weekly_summary_input(weekly_input_id, report_week, staff_name, cost_tracking_id, summary, target, health, user=None):
    _admin_only(user)
    old = get_weekly_summary_by_id(weekly_input_id)
    if not old:
        raise ValueError("找不到 Weekly_Input_ID。")
    if not summary or not summary.strip() or not target or not target.strip():
        raise ValueError("週報內容不可為空。")
    if len(summary) > 500 or len(target) > 500:
        raise ValueError("週報欄位限制 500 字以內。")
    dup = query_df("""
        SELECT Weekly_Input_ID FROM weekly_summary_input
        WHERE Report_Week=? AND Staff_Name=? AND Cost_Tracking_ID=? AND Weekly_Input_ID<>?
    """, [report_week, staff_name, cost_tracking_id, weekly_input_id])
    if not dup.empty:
        raise ValueError("相同 Report_Week + Staff_Name + Cost_Tracking_ID 的週報已存在。")

    execute("""
        UPDATE weekly_summary_input
        SET Report_Week=?, Staff_Name=?, Cost_Tracking_ID=?, Weekly_Summary=?,
            Next_Week_Target=?, Health=?, Updated_At=?, Updated_By=?
        WHERE Weekly_Input_ID=?
    """, [
        report_week, staff_name, cost_tracking_id, summary.strip(), target.strip(),
        health, now(), user["User_Email"], weekly_input_id
    ])
    rebuild_weekly_report_log()
    audit("ADMIN_UPDATE", "weekly_summary_input", weekly_input_id, old, {
        "Report_Week": report_week, "Staff_Name": staff_name, "Cost_Tracking_ID": cost_tracking_id
    }, user["User_Email"])


def delete_weekly_summary_input(weekly_input_id, user=None):
    _admin_only(user)
    old = get_weekly_summary_by_id(weekly_input_id)
    if not old:
        raise ValueError("找不到 Weekly_Input_ID。")
    execute("DELETE FROM weekly_summary_input WHERE Weekly_Input_ID=?", [weekly_input_id])
    rebuild_weekly_report_log()
    audit("ADMIN_DELETE", "weekly_summary_input", weekly_input_id, old, "", user["User_Email"])


# -----------------------------
# Excel Admin Config v3
# -----------------------------

CONFIG_SHEETS = {
    "Staff_Master": ["Staff_Name", "Department", "Role", "Active_Flag"],
    "User_Master": ["User_Email", "Staff_Name", "Department", "Role", "Active_Flag"],
    "Role_Permission": ["Role", "Page_Name", "Can_View", "Can_Edit", "Can_Export"],
    "Master_Lists": ["List_Type", "List_Value", "Active_Flag"],
}

def _style_ws(ws):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 12
        for row in range(1, min(ws.max_row, 200) + 1):
            v = ws[f"{letter}{row}"].value
            if v is not None:
                max_len = max(max_len, min(len(str(v)) + 2, 50))
            ws[f"{letter}{row}"].border = border
            ws[f"{letter}{row}"].alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[letter].width = max_len

def _write_df_to_ws(wb, sheet_name, df):
    ws = wb.create_sheet(sheet_name)
    if df is None or df.empty:
        ws.append(CONFIG_SHEETS.get(sheet_name, ["No Data"]))
    else:
        ws.append(list(df.columns))
        for _, row in df.iterrows():
            ws.append([row[col] for col in df.columns])
    _style_ws(ws)

def export_admin_config_workbook():
    wb = Workbook()
    wb.remove(wb.active)

    _write_df_to_ws(wb, "Staff_Master", get_table("staff_master")[CONFIG_SHEETS["Staff_Master"]])
    _write_df_to_ws(wb, "User_Master", get_table("user_master")[CONFIG_SHEETS["User_Master"]])
    _write_df_to_ws(wb, "Role_Permission", get_table("role_permission")[CONFIG_SHEETS["Role_Permission"]])
    _write_df_to_ws(wb, "Master_Lists", get_table("master_lists")[CONFIG_SHEETS["Master_Lists"]])

    guide = pd.DataFrame([
        {"Item": "Role", "Rule": "Admin / Manager / Staff / PD/PM / Viewer"},
        {"Item": "Page_Name", "Rule": "必須是系統左側選單名稱之一"},
        {"Item": "Active_Flag", "Rule": "只能填 1 或 0"},
        {"Item": "Can_View / Can_Edit / Can_Export", "Rule": "只能填 1 或 0"},
        {"Item": "User_Master.Staff_Name", "Rule": "必須存在於 Staff_Master"},
        {"Item": "Department", "Rule": "必須存在於 Master_Lists 中 List_Type=Department"},
        {"Item": "Import Mode", "Rule": "Staff/User/Master 使用 upsert；Role_Permission 會全量重建"},
    ])
    _write_df_to_ws(wb, "Validation_Guide", guide)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"GLOBAL_Weekly_Report_Admin_Config_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return filename, output.getvalue()

def _normalize_config_df(df, required_cols):
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise ValueError("缺少欄位：" + ", ".join(missing))
    df = df[required_cols]
    df = df.dropna(how="all")
    for col in df.columns:
        if df[col].dtype == "object":
            df[col] = df[col].astype(str).str.strip()
    return df

def read_admin_config_excel(uploaded_file):
    frames = {}
    for sheet, cols in CONFIG_SHEETS.items():
        try:
            raw = pd.read_excel(uploaded_file, sheet_name=sheet)
        except Exception as e:
            raise ValueError(f"無法讀取 Sheet: {sheet}. {e}")
        frames[sheet] = _normalize_config_df(raw, cols)
    return frames

def validate_admin_config_frames(frames):
    errors = []
    summary = []

    for sheet, cols in CONFIG_SHEETS.items():
        if sheet not in frames:
            errors.append(f"缺少 Sheet：{sheet}")
            continue
        df = frames[sheet]
        missing = [c for c in cols if c not in df.columns]
        if missing:
            errors.append(f"{sheet} 缺少欄位：{', '.join(missing)}")
        summary.append({"Sheet": sheet, "Rows": len(df)})

    if errors:
        return errors, pd.DataFrame(summary)

    staff = frames["Staff_Master"]
    users = frames["User_Master"]
    perms = frames["Role_Permission"]
    masters = frames["Master_Lists"]

    def check_binary(df, cols, sheet):
        for col in cols:
            bad = df[~df[col].fillna(0).astype(int).isin([0, 1])]
            if not bad.empty:
                errors.append(f"{sheet}.{col} 只能是 0 或 1。")

    for name, df, key_cols in [
        ("Staff_Master", staff, ["Staff_Name"]),
        ("User_Master", users, ["User_Email"]),
        ("Role_Permission", perms, ["Role", "Page_Name"]),
        ("Master_Lists", masters, ["List_Type", "List_Value"]),
    ]:
        for col in key_cols:
            if df[col].isna().any() or (df[col].astype(str).str.strip() == "").any():
                errors.append(f"{name}.{col} 不可空白。")
        dup = df[df.duplicated(key_cols, keep=False)]
        if not dup.empty:
            errors.append(f"{name} 主鍵重複：{key_cols}")

    check_binary(staff, ["Active_Flag"], "Staff_Master")
    check_binary(users, ["Active_Flag"], "User_Master")
    check_binary(perms, ["Can_View", "Can_Edit", "Can_Export"], "Role_Permission")
    check_binary(masters, ["Active_Flag"], "Master_Lists")

    allowed_depts = set(masters[(masters["List_Type"] == "Department") & (masters["Active_Flag"].astype(int) == 1)]["List_Value"].astype(str))
    allowed_roles = set(ROLES)
    allowed_pages = set(PAGE_ORDER)
    staff_names = set(staff["Staff_Name"].astype(str))

    for dept in staff["Department"].astype(str):
        if dept not in allowed_depts:
            errors.append(f"Staff_Master.Department 不存在於 Master_Lists Department：{dept}")

    for _, row in users.iterrows():
        if str(row["Staff_Name"]) not in staff_names:
            errors.append(f"User_Master.Staff_Name 不存在於 Staff_Master：{row['Staff_Name']}")
        if str(row["Department"]) not in allowed_depts:
            errors.append(f"User_Master.Department 不存在於 Master_Lists Department：{row['Department']}")
        if str(row["Role"]) not in allowed_roles:
            errors.append(f"User_Master.Role 不合法：{row['Role']}")

    for _, row in perms.iterrows():
        if str(row["Role"]) not in allowed_roles:
            errors.append(f"Role_Permission.Role 不合法：{row['Role']}")
        if str(row["Page_Name"]) not in allowed_pages:
            errors.append(f"Role_Permission.Page_Name 不存在：{row['Page_Name']}")

    for role in ROLES:
        if role not in set(perms["Role"].astype(str)):
            errors.append(f"Role_Permission 缺少角色設定：{role}")

    return errors, pd.DataFrame(summary)

def apply_admin_config_frames(frames, user):
    _admin_only(user)
    errors, summary = validate_admin_config_frames(frames)
    if errors:
        raise ValueError("Validation failed: " + " / ".join(errors[:10]))

    user_email = user["User_Email"]
    n = now()

    staff = frames["Staff_Master"].copy()
    users = frames["User_Master"].copy()
    perms = frames["Role_Permission"].copy()
    masters = frames["Master_Lists"].copy()

    with get_connection() as conn:
        c = conn.cursor()

        for _, row in masters.iterrows():
            c.execute("""
                INSERT INTO master_lists (List_Type, List_Value, Active_Flag)
                VALUES (?, ?, ?)
                ON CONFLICT(List_Type, List_Value)
                DO UPDATE SET Active_Flag=excluded.Active_Flag
            """, [str(row["List_Type"]), str(row["List_Value"]), int(row["Active_Flag"])])

        for _, row in staff.iterrows():
            c.execute("""
                INSERT INTO staff_master (Staff_Name, Department, Role, Active_Flag)
                VALUES (?, ?, ?, ?)
                ON CONFLICT(Staff_Name)
                DO UPDATE SET Department=excluded.Department, Role=excluded.Role, Active_Flag=excluded.Active_Flag
            """, [str(row["Staff_Name"]), str(row["Department"]), str(row["Role"]), int(row["Active_Flag"])])

        for _, row in users.iterrows():
            c.execute("""
                INSERT INTO user_master (User_Email, Staff_Name, Department, Role, Active_Flag, Created_At, Updated_At)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(User_Email)
                DO UPDATE SET Staff_Name=excluded.Staff_Name,
                              Department=excluded.Department,
                              Role=excluded.Role,
                              Active_Flag=excluded.Active_Flag,
                              Updated_At=excluded.Updated_At
            """, [str(row["User_Email"]), str(row["Staff_Name"]), str(row["Department"]),
                  str(row["Role"]), int(row["Active_Flag"]), n, n])

        c.execute("DELETE FROM role_permission")
        for _, row in perms.iterrows():
            c.execute("""
                INSERT INTO role_permission (Role, Page_Name, Can_View, Can_Edit, Can_Export)
                VALUES (?, ?, ?, ?, ?)
            """, [str(row["Role"]), str(row["Page_Name"]), int(row["Can_View"]),
                  int(row["Can_Edit"]), int(row["Can_Export"])])

        conn.commit()

    audit("ADMIN_CONFIG_IMPORT", "admin_config_excel", "GLOBAL_Weekly_Report_Admin_Config", "", {
        "Staff_Master": len(staff),
        "User_Master": len(users),
        "Role_Permission": len(perms),
        "Master_Lists": len(masters),
    }, user_email)
    return summary
